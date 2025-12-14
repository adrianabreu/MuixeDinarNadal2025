import openpyxl
import json
import shutil

# Carregar l'Excel
wb = openpyxl.load_workbook('../barret_magic_muixeranga_COMPLET.xlsx')
ws = wb.active

# Obtenir les capçaleres
headers = [cell.value for cell in ws[1]]
print(f"Capçaleres trobades: {headers}")

# Mapa de normalització de categories
def normalize_category(cat):
    if not cat:
        return cat
    cat_lower = cat.lower()
    if 'muixe' in cat_lower:
        return 'Muixelovers'
    elif 'fomo' in cat_lower:
        return 'FOMO de Ferro'
    elif 'comboiet' in cat_lower:
        return 'Comboiet'
    elif 'talent' in cat_lower:
        return 'Talents Emergents'
    return cat

# Extreure les dades
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # Si té nom
        member = {
            'nom': row[0],
            'alias': row[1],
            'context': row[2],
            'fraseBarret': row[3],
            'categoria': normalize_category(row[4])
        }
        data.append(member)

# Guardar com a JSON a public/
with open('public/members.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# Copiar també a src/assets/ per si es necessita
shutil.copy('public/members.json', 'src/assets/members.json')

print(f"\n✅ Convertit {len(data)} membres")
print(f"   📁 public/members.json")
print(f"   📁 src/assets/members.json")
print(f"\nCategories trobades:")
categories = set(m['categoria'] for m in data if m['categoria'])
for cat in sorted(categories):
    count = len([m for m in data if m['categoria'] == cat])
    print(f"  - {cat}: {count} membres")

print(f"\n💡 Refresca el navegador amb Cmd+Shift+R (Mac) o Ctrl+Shift+R (Win/Linux)")


